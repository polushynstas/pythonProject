from person import Person
import openpyxl

class PeopleDatabase:
    def __init__(self):
        self.people = []
        self.load_from_file()

    def add_person(self, person):
        self.people.append(person)

    def search_people(self, query):
        result = []
        for person in self.people:
            if query.lower() in person.first_name.lower() or query.lower() in person.last_name.lower() or query.lower() in person.middle_name.lower():
                result.append(person)
        return result

    def save_to_file(self, filename="dataset.xlsx"):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        headers = ["Ім'я", "Прізвище", "По-батькові", "Дата народження", "Дата смерті", "Стать", "Вік"]
        sheet.append(headers)

        for person in self.people:
            row_data = [
                person.first_name,
                person.last_name,
                person.middle_name,
                person.birth_date.strftime('%d.%m.%Y') if person.birth_date else '',
                person.death_date.strftime('%d.%m.%Y') if person.death_date else '',
                person.gender,
                person.calculate_age() if person.calculate_age() is not None else ''
            ]
            sheet.append(row_data)
        try:
            workbook.save(filename)
            print(f"Данні збережено у файлі {filename}.")
        except Exception as e:
            print(f"Помилка при збереженні у файл: {e}")


    def load_from_file(self, filename="dataset.xlsx"):
        try:
            workbook = openpyxl.load_workbook(filename)
            sheet = workbook.active

            self.people = []
            for row_num in range(2, sheet.max_row + 1):
                first_name = sheet.cell(row=row_num, column=1).value
                last_name = sheet.cell(row=row_num, column=2).value
                middle_name = sheet.cell(row=row_num, column=3).value
                birth_date = sheet.cell(row=row_num, column=4).value
                death_date = sheet.cell(row=row_num, column=5).value
                gender = sheet.cell(row=row_num, column=6).value

                person = Person(first_name, last_name, middle_name, birth_date, death_date, gender)
                self.people.append(person)

            print(f"Данні завантажено з файлу {filename}.")

        except FileNotFoundError:
            print(f"Файл {filename} не знайдено. Нова база даних створена.")

if __name__ == "__main__":
    database = PeopleDatabase()

    while True:
        print("\n1. Додати нову людину")
        print("2. Пошук людей")
        print("3. Завантажити дані з файлу")
        print("4. Вийти")

        choice = input("Оберіть опцію: ")

        if choice == "1":
            first_name = input("Введіть ім'я: ")
            last_name = input("Введіть прізвище: ")
            middle_name = input("Введіть по-батькові: ")
            birth_date = input("Введіть дату народження (формат: дд.мм.рррр): ")
            death_date = input("Введіть дату смерті (або натисніть Enter, якщо немає): ")
            gender = input("Введіть стать (Ч/Ж): ")

            person = Person(first_name, last_name, middle_name, birth_date, death_date, gender)
            database.add_person(person)
            print("Людина додана до бази.")
            database.save_to_file()

        elif choice == "2":
            query = input("Введіть ім'я, прізвище або по-батькові для пошуку: ")
            results = database.search_people(query)
            if results:
                for person in results:
                    person.display_info()
            else:
                print("Збігів не знайдено.")


        elif choice == "3":
            filename = input("Введіть ім'я файлу для завантаження даних: ")
            database.load_from_file(filename)

        elif choice == "4":
            break

        else:
            print("Неправильний вибір. Спробуйте ще раз.")